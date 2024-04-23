# -*- coding : utf-8 -*-
# -*- author : zoro123 -*-
# -*- version : v2.3 -*-
# -*- date : 2024.04 -*-

import re
import time
import openpyxl
import sys
import os
import chardet
import openpyxl as p

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from chardet.universaldetector import UniversalDetector



def get_encoding(file):
    # 二进制方式读取，获取字节数据，检测类型
    with open(file, 'rb') as f:
        data = f.read()
        return chardet.detect(data)['encoding']

def get_encode_info(file):
    with open(file, 'rb') as f:
        data = f.read()
        result = chardet.detect(data)
        return result['encoding']

def read_file(file):
    with open(file, 'rb') as f:
        return f.read()

def write_file(content, file):
    with open(file, 'wb') as f:
        f.write(content)


def convert_encode2utf8(file, original_encode, des_encode):
    file_content = read_file(file)
    file_decode = file_content.decode(original_encode, 'ignore')
    file_encode = file_decode.encode(des_encode)
    write_file(file_encode, file)


def OpenFile():

    file_name = getInput()
    datalist = []
    datastr = ''

    encode_info = get_encode_info(file_name)

    if encode_info == 'utf-8':

        with open(file_name, encoding='utf-8') as f:
            for i in f.readlines():
                datalist.append(i.strip())
        with open(file_name, encoding='utf-8') as f:
            datastr = f.read()

    elif encode_info != 'utf-8':

        convert_encode2utf8(file_name, encode_info, 'utf-8')

        with open(file_name, encoding='utf-8') as f:
            for i in f.readlines():
                datalist.append(i.strip())
        with open(file_name, encoding='utf-8') as f:
            datastr = f.read()

    return datalist, datastr

#输出存活端口
def OpenPort(datalist):

    sheetList = [['ip', 'port']]

    for i in datalist:
        p = re.findall(r'^\d[^\s]+', i)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                # exit()
                # port = u.replace(ip[0], '').strip(':')
                port = re.findall("(?<=:)\d+" , u)
                # exit()
                try:
                    # exit()
                    ip.append(port[0])
                    sheetList.append(ip)
                except IndexError:
                    pass

    OutPut('OpenPort', sheetList)

#输出IP段内存货数量
def AliveIp(datalist):

    sheetList = [['IP range', 'Active IP ranges']]


    for t in datalist:
        Ip_d = re.findall(r"\[\*]\sLiveTop\s\d+\.\d+\.\d+\.\d+/\d+.*", t)

        if len(Ip_d) != 0:
            p1 = list(Ip_d)

            for u in p1:
                ip_duan = re.findall(r"\d+\.\d+\.\d+\.\d+/\d+", u)
                No = re.findall(r"\d+$", u)
                ip_duan.append(No[0])
                sheetList.append(ip_duan)

    OutPut('AliveIp', sheetList)


#输出识别到的系统
def Oslist(datalist):

    replaceList = ["[*]", '\t', "\x01", '\x02']

    sheetList = [['ip', 'os']]

    for t in datalist:
        p = re.findall(r"\[\*]\s\d+\.\d+\.\d+\.\d+.*", t)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                #删除无用字符
                for q in replaceList:
                    u = u.replace(q, "")

                ip.append(u.replace(ip[0], '').strip())
                sheetList.append(ip)

    OutPut('OsList', sheetList)

#输出exp漏洞列表
def Bug_ExpList(datalist):

    sheetList = [['ip', 'bug_exp']]

    for i in datalist:
        p = re.findall(r"\[\+]\s\d+\.\d+\.\d+\.\d+.*", i)

        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                bug = u.replace(ip[0], '').replace("[+]", "").replace('\t', '').strip()
                ip.append(bug)
                sheetList.append(ip)

    OutPut('Bug_ExpList', sheetList)

#输出poc漏洞列表
def Bug_PocList(datalist):

    sheetList = [['url', 'bug_poc']]

    for i in datalist:
        p = re.findall(r"\[\+]\shttp[^\s].*", i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                # url = re.findall(r"http[^\s].*\s", u)
                url = re.findall(r"(?P<url>https?://\S+)", u)
                bug = u.replace(url[0], '').replace("[+]", "").replace('\t', '').strip()
                url.append(bug)
                sheetList.append(url)

    OutPut('Bug_PocList', sheetList)

#输出title
def GetTitle(datalist):

    sheetList = [['url', 'code', 'len', 'title']]

    for i in datalist:
        p = re.findall(r'\[\*]\sWebTitle.*', i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                all_url = re.findall(r"http[^\s]+", u)
                url = [all_url[0]]
                code = re.findall(r'(?<=code:)[^\s]+', u)
                len1 = re.findall(r'(?<=len:)[^\s]+', u)
                title = re.findall(r'(?<=title:).*', u)

                # url.append(str(code).strip("['").strip("']'"))
                url.append(code[0])
                url.append(str(len1).strip("['").strip("']'"))
                url.append(str(title).strip("['").strip("']'"))
                # print(url)
                # exit()
                sheetList.append(url)

    OutPut('Title', sheetList)

#输出弱口令
def GetPassword(datalist):

    sheetList = [['ip', 'port', 'server', 'user&passwd']]

    for i in datalist:
        p = re.findall(r'((ftp|mysql|mssql|SMB|RDP|Postgres|SSH|oracle|SMB2-shares)(:|\s).*)', i, re.I)
        rd = re.findall(r'((redis|Mongodb)(:|\s).*)', i, re.I)
        mc = re.findall(r"((Memcached)(:|\s).*)", i, re.I)

        if len(p) != 0 and p[0][-1] == ":":
            p1 = list(p)

            all = p1[0][0].split(":")
            try:
                passwd = all[3]
            except:
                passwd = []
                pass
            server = all[0]
            port = all[2]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", str(all))
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)


        if len(rd) != 0 and len(rd[0][0].split(" ")) == 2:
            rd1 = list(rd)

            rd_all = rd1[0][0].split(" ")
            passwd = rd_all[-1]
            server = rd1[0][1]
            port = (rd_all[0].split(":"))[2]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", rd1[0][0])
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

        if len(mc) != 0:
            mc1 = list(mc)

            mc_all = mc1[0][0].split(" ")
            passwd = mc_all[2]
            server = mc_all[0]
            port = (mc_all[1].split(":"))[-1]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", mc1[0][0])
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

    OutPut('WeakPasswd', sheetList)


#输出指纹信息
def FingerOut(datalist):

    # w1 = wb.create_sheet('')
    sheetList = [['url', 'finger']]

    for i in datalist:
        p = re.findall(r'.*InfoScan.*', i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r'http[^\s]+', u)
                finger = u.split(url[0])[-1].strip()
                url.append(finger)
                # ws4.append(url)
                sheetList.append(url)

    OutPut('Finger', sheetList)

#表格输出整理
def OutPut(sheetname,sheetList):

    sheetName = wb.create_sheet(sheetname)



    input_filename = sys.argv[1].split(".txt")[0]

    print(f"[*]{input_filename}.txt读取成功，{input_filename}_{sheetname}.txt文件生成中······\n")

    delimiter = "================{}================\n"
    with open(f"{input_filename}_{sheetname}.txt", 'w', encoding='utf-8') as txt_file:
        txt_file.write(delimiter.format(sheetname))
        for row in sheetList:
            txt_file.write("\t".join(map(str, row)) + "\n")

    #将列表写入sheet
    for i in sheetList:
        # 解决\x03此类特殊字符报错
        try:
            sheetName.append(i)
        except openpyxl.utils.exceptions.IllegalCharacterError:
            i[-1] = ILLEGAL_CHARACTERS_RE.sub(r'', i[-1])
            sheetName.append(i)
        except Exception as e:
            print(f"err: {e}")


    #首行格式
    for row in sheetName[f"A1:{chr(65 + len(list1[0]) - 1)}1"]:
        for cell in row:
            cell.font = Font(size=12, bold=True)


def getInput():

    if len(sys.argv) != 2:
        print("\n[*] fscan结果整理脚本，输出为.xlsx文件\n\nUsage: \n\n    python3 FscanOutput.py result.txt\n")
        exit()

    if not os.path.exists(sys.argv[1]): 
        print(f"[{sys.argv[1]}] 文件不存在")
        exit()

    return sys.argv[1]

# def write_to_txt(data, filename="res.txt"):
#     with open(filename, 'a', encoding='utf-8',errors='ignore') as f:
#         for item in data:
#             f.write("%s\n" % item)


if __name__ == "__main__":

    print(r'''

============================================================
 ______                    ____        _               _   
|  ____|                  / __ \      | |             | |  
| |__ ___  ___ __ _ _ __ | |  | |_   _| |_ _ __  _   _| |_ 
|  __/ __|/ __/ _` | '_ \| |  | | | | | __| '_ \| | | | __|
| |  \__ \ (_| (_| | | | | |__| | |_| | |_| |_) | |_| | |_ 
|_|  |___/\___\__,_|_| |_|\____/ \__,_|\__| .__/ \__,_|\__|
                                          | |              
                                          |_|           Pro       
============================================================
                                         ---By zoro123 v2.3
           ''')
    list1, str1 = OpenFile()

    wb = openpyxl.Workbook()
    OpenPort(list1)
    AliveIp(list1)
    Bug_ExpList(list1)
    Bug_PocList(list1)
    Oslist(list1)
    GetTitle(list1)
    GetPassword(list1)
    FingerOut(list1)
    ws5 = wb["Sheet"]
    wb.remove(ws5)
    input_filename = sys.argv[1].split(".txt")[0]
    Output_xlsx = (f"%s_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx" % input_filename)
    wb.save(Output_xlsx)
    print("[+]文件处理结果如下······\n")
    New_fscanxlsx = p.load_workbook(Output_xlsx)

    print("+---------------------------------+\n")
    wt = New_fscanxlsx['OpenPort']
    print("[+++]探测存活端口共计：%s 个" % (wt.max_row-1))
    wt = New_fscanxlsx['AliveIp']
    print("[+++]探测存活IP段共计：%s 个" % (wt.max_row - 1))
    wt1 = New_fscanxlsx['Bug_ExpList']
    print("[+++]Exp可利用漏洞共计：%s 个" % (wt1.max_row-1))
    wt2 = New_fscanxlsx['Bug_PocList']
    print("[+++]Poc可利用漏洞共计：%s 个" % (wt2.max_row-1))
    wt3 = New_fscanxlsx['OsList']
    print("[+++]成功识别操作系统共计：%s 个" % (wt3.max_row-1))
    wt4 = New_fscanxlsx['Title']
    print("[+++]成功探测Web服务共计：%s 条" % (wt4.max_row-1))
    wt5 = New_fscanxlsx['WeakPasswd']
    print("[+++]成功破解账号密码共计：%s 个" % (wt5.max_row-1))
    wt6 = New_fscanxlsx['Finger']
    print("[+++]成功识别指纹共计：%s 个\n" % (wt6.max_row-1))
    print("+---------------------------------+\n")

    print('[+]结果已经整理输出至 -- %s -- 文件所在目录！\n' % sys.argv[1])
    print('--> 表格文件名为：%s\n' % Output_xlsx)
    print(f'--> 各个模块处理文件名为：{input_filename}_sheetName.txt')










