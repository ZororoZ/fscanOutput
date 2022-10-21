# -*- coding : utf-8 -*-
# -*- author : zoro123 -*-
# -*- version : v2.0 -*-
# -*- date : 2022 -*-

import re
import time
import openpyxl
import sys
import os
import openpyxl as p

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font


def OpenFile():

    filename = getInput()
    datalist = []
    datastr = ''

    with open(filename, encoding='utf-8') as f:
        for i in f.readlines():
            datalist.append(i.strip())

    with open(filename, encoding='utf-8') as f:
        datastr = f.read()

    return datalist, datastr

#输出存活端口
def OpenPort(datalist):

    sheetList = [['ip', 'port']]

    for i in datalist:
        p = re.findall(r'^\d[^\s]+', i)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                # port = u.replace(ip[0], '').strip(':')
                port = re.findall("(?<=:)\d+" , u)
                # print(port)
                # exit()
                ip.append(port[0])
                sheetList.append(ip)

    OutPut('OpenPort', sheetList)

#输出识别到的系统
def Oslist(datalist):

    replaceList = ["[*]", '\t', "\x01", '\x02']

    sheetList = [['ip', 'os']]

    for t in datalist:
        p = re.findall(r"\[\*]\s\d+\.\d+\.\d+\.\d+.*", t)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                #删除无用字符
                for q in replaceList:
                    u = u.replace(q, "")

                ip.append(u.replace(ip[0], '').strip())
                sheetList.append(ip)

    OutPut('OsList', sheetList)

#输出exp漏洞列表
def Bug_ExpList(datalist):

    sheetList = [['ip', 'bug_exp']]

    for i in datalist:
        p = re.findall(r"\[\+]\s\d+\.\d+\.\d+\.\d+.*", i)

        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                bug = u.replace(ip[0], '').replace("[+]", "").replace('\t', '').strip()
                ip.append(bug)
                sheetList.append(ip)

    OutPut('Bug_ExpList', sheetList)

#输出poc漏洞列表
def Bug_PocList(datalist):

    sheetList = [['url', 'bug_poc']]

    for i in datalist:
        p = re.findall(r"\[\+]\shttp[^\s].*", i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r"http[^\s].*\s", u)
                bug = u.replace(url[0], '').replace("[+]", "").replace('\t', '').strip()
                url.append(bug)
                sheetList.append(url)

    OutPut('Bug_PocList', sheetList)

#输出title
def GetTitle(datalist):

    sheetList = [['url', 'code', 'len', 'title']]

    for i in datalist:
        p = re.findall(r'\[\*]\sWebTitle.*', i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r"http[^\s]+", u)
                code = re.findall(r'(?<=code:)[^\s]+', u)
                len1 = re.findall(r'(?<=len:)[^\s]+', u)
                title = re.findall(r'(?<=title:).*', u)
                # print(title)
                url.append(str(code).strip("['").strip("']'"))
                url.append(str(len1).strip("['").strip("']'"))
                url.append(str(title).strip("['").strip("']'"))
                sheetList.append(url)

    OutPut('Title', sheetList)

#输出弱口令
def GetPassword(datalist):

    sheetList = [['ip', 'server', 'passwd']]

    for i in datalist:
        p = re.findall(r'((ftp|mysql|mssql|SMB|RDP|Postgres|SSH|Mongodb|oracle|redis|Memcached)(:|\s).*)', i, re.I)

        if len(p) != 0:
            p1 = list(p)

            passwd = p1[0][0]
            server = p1[0][1]

            ip = re.findall(r"\d+\.\d+\.\d+\.\d+\:\d+", passwd)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

    OutPut('WeakPasswd', sheetList)


#输出指纹信息
def FingerOut(datalist):

    # w1 = wb.create_sheet('')
    sheetList = [['url', 'finger']]

    for i in datalist:
        p = re.findall(r'.*InfoScan.*', i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r'http[^\s]+', u)
                finger = u.split(url[0])[-1].strip()
                url.append(finger)
                # ws4.append(url)
                sheetList.append(url)

    OutPut('Finger', sheetList)

#表格输出整理
def OutPut(sheetname,sheetList):

    sheetName = wb.create_sheet(sheetname)

    #将列表写入sheet
    for i in sheetList:
        # 解决\x03此类特殊字符报错
        try:
            sheetName.append(i)
        except openpyxl.utils.exceptions.IllegalCharacterError:
            i[-1] = ILLEGAL_CHARACTERS_RE.sub(r'', i[-1])
            sheetName.append(i)
        except Exception as e:
            print(f"err: {e}")


    #首行格式
    for row in sheetName[f"A1:{chr(65 + len(list1[0]) - 1)}1"]:
        for cell in row:
            cell.font = Font(size=12, bold=True)


def getInput():

    if len(sys.argv) != 2:
        print("\nfscan结果整理脚本，输出为.xlsx文件\n\nUsage: \n    python3 FscanOutput_v1.02.py result.txt\n")
        exit()

    if not os.path.exists(sys.argv[1]):
        print(f"[{sys.argv[1]}] 文件不存在")
        exit()

    return sys.argv[1]

if __name__ == "__main__":

    print(r'''
 ______                    ____        _               _   
|  ____|                  / __ \      | |             | |  
| |__ ___  ___ __ _ _ __ | |  | |_   _| |_ _ __  _   _| |_ 
|  __/ __|/ __/ _` | '_ \| |  | | | | | __| '_ \| | | | __|
| |  \__ \ (_| (_| | | | | |__| | |_| | |_| |_) | |_| | |_ 
|_|  |___/\___\__,_|_| |_|\____/ \__,_|\__| .__/ \__,_|\__|
                                          | |              
                                          |_|             
============================================================
                                         ---By zoro123 v2.0
           ''')
    list1, str1 = OpenFile()

    wb = openpyxl.Workbook()
    OpenPort(list1)
    Bug_ExpList(list1)
    Bug_PocList(list1)
    Oslist(list1)
    GetTitle(list1)
    GetPassword(list1)
    FingerOut(list1)
    ws5 = wb["Sheet"]
    wb.remove(ws5)
    input_filename = sys.argv[1].split(".txt")[0]
    Output_xlsx = (f"%s_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx" % input_filename)
    # wb.save(f"%s_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx" %input_filename)
    wb.save(Output_xlsx)
    print("[+]文件读取成功，处理结果如下······\n")
    New_fscanxlsx = p.load_workbook(Output_xlsx)

    print("+---------------------------------+\n")
    wt = New_fscanxlsx['OpenPort']
    # print(type(wt.max_row))
    # exit()
    print("[+++]探测存活端口共计：%s 个" % (wt.max_row-1))
    wt1 = New_fscanxlsx['Bug_ExpList']
    print("[+++]Exp可利用漏洞共计：%s 个" % (wt1.max_row-1))
    wt2 = New_fscanxlsx['Bug_PocList']
    print("[+++]Poc可利用漏洞共计：%s 个" % (wt2.max_row-1))
    wt3 = New_fscanxlsx['OsList']
    print("[+++]成功识别操作系统共计：%s 个" % (wt3.max_row-1))
    wt4 = New_fscanxlsx['Title']
    print("[+++]成功探测Web服务共计：%s 条" % (wt4.max_row-1))
    wt5 = New_fscanxlsx['WeakPasswd']
    print("[+++]成功破解账号密码共计：%s 个" % (wt5.max_row-1))
    wt6 = New_fscanxlsx['Finger']
    print("[+++]成功识别指纹共计：%s 个" % (wt6.max_row-1))
    print("+---------------------------------+\n")

    print('[+]结果已经整理输出至 -- %s -- 文件所在目录！\n' % sys.argv[1])
    print('--> 文件名为：%s' % Output_xlsx)










