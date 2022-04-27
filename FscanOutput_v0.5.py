# -*- coding : utf-8 -*-
# -*- author : zoro123 -*-
# -*- version : v0.5 -*-
# -*- date : 2022 -*-

import re
import time
import openpyxl
import sys
import os
from openpyxl.styles import Font


def OpenFile():

    filename = getInput()
    datalist = []
    datastr = ''

    with open(filename, encoding='utf-8') as f:
        for i in f.readlines():
            datalist.append(i.strip())

    with open(filename, encoding='utf-8') as f:
        datastr = f.read()

    return datalist, datastr

#输出存活端口
def OpenPort(datalist):

    sheetList = [['ip', 'port']]

    for i in datalist:
        p = re.findall(r'^\d[^\s]+', i)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                port = u.replace(ip[0], '').strip(':')
                ip.append(port)
                sheetList.append(ip)

    OutPut('OpenPort', sheetList)

#输出识别到的系统
def Oslist(datalist):

    replaceList = ["[*]", '\t', "\x01", '\x02']

    sheetList = [['ip', 'os']]

    for t in datalist:
        p = re.findall(r"\[\*]\s\d+\.\d+\.\d+\.\d+.*", t)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                #删除无用字符
                for q in replaceList:
                    u = u.replace(q, "")

                ip.append(u.replace(ip[0], '').strip())
                sheetList.append(ip)

    OutPut('OsList', sheetList)

#输出漏洞列表
def BugList(datalist):

    sheetList = [['ip', 'bug']]

    for i in datalist:
        p = re.findall(r"\[\+]\s\d+\.\d+\.\d+\.\d+.*", i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                bug = u.replace(ip[0], '').replace("[+]", "").replace('\t', '').strip()
                ip.append(bug)
                sheetList.append(ip)

    OutPut('BugList', sheetList)

#输出title
def GetTitle(datalist):

    sheetList = [['url', 'code', 'len', 'title']]

    for i in datalist:
        p = re.findall(r'\[\*]\sWebTitle.*', i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r"http[^\s]+", u)
                code = re.findall(r'(?<=code:)[^\s]+', u)
                len1 = re.findall(r'(?<=len:)[^\s]+', u)
                title = re.findall(r'(?<=title:).*', u)
                # print(title)
                url.append(str(code).strip("['").strip("']'"))
                url.append(str(len1).strip("['").strip("']'"))
                url.append(str(title).strip("['").strip("']'"))
                sheetList.append(url)

    OutPut('Title', sheetList)

#输出弱口令
def GetPassword(datalist):

    sheetList = [['ip', 'server', 'passwd']]

    for i in datalist:
        p = re.findall(r'((ftp|mysql|mssql|SMB|RDP|Postgres|SSH|Mongodb|oracle|redis):.*)', i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            passwd = p1[0][0]
            server = p1[0][1]
            # print(passwd)
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+\:\d+", passwd)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

    OutPut('WeakPasswd', sheetList)


#输出指纹信息
def FingerOut(datalist):

    # w1 = wb.create_sheet('')
    sheetList = [['url', 'finger']]

    for i in datalist:
        p = re.findall(r'.*InfoScan.*', i)
        # print(p)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r'http[^\s]+', u)
                finger = u.split(url[0])[-1].strip()
                url.append(finger)
                # ws4.append(url)
                sheetList.append(url)

    OutPut('Finger', sheetList)

#表格输出整理
def OutPut(sheetname,sheetList):

    sheetName = wb.create_sheet(sheetname)

    #将列表写入sheet
    for i in sheetList:
        # print(i)
        sheetName.append(i)


    #首行格式
    for row in sheetName[f"A1:{chr(65 + len(list1[0]) - 1)}1"]:
        for cell in row:
            cell.font = Font(size=12, bold=True)


def getInput():

    if len(sys.argv) != 2:
        print("\nfscan结果整理脚本，输出为.xlsx文件\n\nUsage: \n    python3 FscanOutput_v0.5.py result.txt\n")
        exit()

    if not os.path.exists(sys.argv[1]):
        print(f"[{sys.argv[1]}] 文件不存在")
        exit()

    return sys.argv[1]

if __name__ == "__main__":

    print(r'''
 ______                    ____        _               _   
|  ____|                  / __ \      | |             | |  
| |__ ___  ___ __ _ _ __ | |  | |_   _| |_ _ __  _   _| |_ 
|  __/ __|/ __/ _` | '_ \| |  | | | | | __| '_ \| | | | __|
| |  \__ \ (_| (_| | | | | |__| | |_| | |_| |_) | |_| | |_ 
|_|  |___/\___\__,_|_| |_|\____/ \__,_|\__| .__/ \__,_|\__|
                                          | |              
                                          |_|             
============================================================
                                               ---By zoro123
           ''')
    list1, str1 = OpenFile()

    wb = openpyxl.Workbook()
    OpenPort(list1)
    BugList(list1)
    Oslist(list1)
    GetTitle(list1)
    GetPassword(list1)
    FingerOut(list1)
    ws5 = wb["Sheet"]
    wb.remove(ws5)
    wb.save(f"fscanResult_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx")

    print('结果已经整理输出至当前目录！')




