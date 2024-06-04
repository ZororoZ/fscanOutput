# -*- coding : utf-8 -*-
# -*- author : zoro123 -*-
# -*- version : v2.3.1 -*-
# -*- date : 2024.06 -*-

import re
import time
import openpyxl
import sys
import os
import chardet
import openpyxl as p

from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from chardet.universaldetector import UniversalDetector



def get_encoding(file):
    # 二进制方式读取，获取字节数据，检测类型
    with open(file, 'rb') as f:
        data = f.read()
        return chardet.detect(data)['encoding']

def get_encode_info(file):
    with open(file, 'rb') as f:
        data = f.read()
        result = chardet.detect(data)
        return result['encoding']

def read_file(file):
    with open(file, 'rb') as f:
        return f.read()

def write_file(content, file):
    with open(file, 'wb') as f:
        f.write(content)


def convert_encode2utf8(file, original_encode, des_encode):
    file_content = read_file(file)
    file_decode = file_content.decode(original_encode, 'ignore')
    file_encode = file_decode.encode(des_encode)
    write_file(file_encode, file)


def OpenFile():

    file_name = getInput()
    datalist = []
    datastr = ''

    encode_info = get_encode_info(file_name)

    if encode_info == 'utf-8':

        with open(file_name, encoding='utf-8') as f:
            for i in f.readlines():
                datalist.append(i.strip())
        with open(file_name, encoding='utf-8') as f:
            datastr = f.read()

    elif encode_info != 'utf-8':

        convert_encode2utf8(file_name, encode_info, 'utf-8')

        with open(file_name, encoding='utf-8') as f:
            for i in f.readlines():
                datalist.append(i.strip())
        with open(file_name, encoding='utf-8') as f:
            datastr = f.read()

    return datalist, datastr

#输出存活端口
def OpenPort(datalist):

    sheetList = [['ip', 'port']]

    for i in datalist:
        p = re.findall(r'^\d[^\s]+', i)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                # exit()
                # port = u.replace(ip[0], '').strip(':')
                port = re.findall("(?<=:)\d+" , u)
                # exit()
                try:
                    # exit()
                    ip.append(port[0])
                    sheetList.append(ip)
                except IndexError:
                    pass

    OutPut('OpenPort', sheetList)

#输出IP段内存货数量
def AliveIp(datalist):

    sheetList = [['IP range', 'Active IP ranges']]


    for t in datalist:
        Ip_d = re.findall(r"\[\*]\sLiveTop\s\d+\.\d+\.\d+\.\d+/\d+.*", t)

        if len(Ip_d) != 0:
            p1 = list(Ip_d)

            for u in p1:
                ip_duan = re.findall(r"\d+\.\d+\.\d+\.\d+/\d+", u)
                No = re.findall(r"\d+$", u)
                ip_duan.append(No[0])
                sheetList.append(ip_duan)

    OutPut('AliveIp', sheetList)


#输出识别到的系统
def Oslist(datalist):

    replaceList = ["[*]", '\t', "\x01", '\x02']

    sheetList = [['ip', 'os']]

    for t in datalist:
        p = re.findall(r"\[\*]\s\d+\.\d+\.\d+\.\d+.*", t)

        if len(p) != 0:
            p1 = list(p)

            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                #删除无用字符
                for q in replaceList:
                    u = u.replace(q, "")

                ip.append(u.replace(ip[0], '').strip())
                sheetList.append(ip)

    OutPut('OsList', sheetList)

#输出exp漏洞列表
def Bug_ExpList(datalist):

    sheetList = [['ip', 'bug_exp']]

    for i in datalist:
        p = re.findall(r"\[\+]\s\d+\.\d+\.\d+\.\d+.*", i)
        new_p = re.findall(r"\[\+]\s\w+-.*", i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u)
                bug = u.replace(ip[0], '').replace("[+]", "").replace('\t', '').strip()
                ip.append(bug)
                sheetList.append(ip)

        if len(new_p) != 0:
            p2 = list(new_p)
            for u1 in p2:
                ip = re.findall(r"\d+\.\d+\.\d+\.\d+", u1)
                bug = u1.replace(ip[0], '').replace("[+]", "").strip()
                ip.append(bug)
                sheetList.append(ip)

    OutPut('Bug_ExpList', sheetList)

#输出poc漏洞列表
def Bug_PocList(datalist):


    sheetList = [['url', 'bug_poc']]

    for i in datalist:
        p = re.findall(r"\[\+].*poc-yaml[^\s].*", i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r"(?P<url>https?://\S+)", u)
                bug = re.findall(r"poc-yaml.*", u)
                url.append(bug[0])
                sheetList.append(url)

    OutPut('Bug_PocList', sheetList)

#输出title
def GetTitle(datalist):

    sheetList = [['url', 'code', 'len', 'title']]

    for i in datalist:
        p = re.findall(r'\[\*]\sWebTitle.*', i)

        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                all_url = re.findall(r"http[^\s]+", u)
                url = [all_url[0]]
                code = re.findall(r'(?<=code:)[^\s]+', u)
                len1 = re.findall(r'(?<=len:)[^\s]+', u)
                title = re.findall(r'(?<=title:).*', u)

                # url.append(str(code).strip("['").strip("']'"))
                url.append(code[0])
                url.append(str(len1).strip("['").strip("']'"))
                url.append(str(title).strip("['").strip("']'"))
                # print(url)
                # exit()
                sheetList.append(url)

    OutPut('Title', sheetList)

#输出弱口令
def GetPassword(datalist):

    sheetList = [['ip', 'port', 'server', 'user&passwd']]

    for i in datalist:
        p = re.findall(r'((ftp|mysql|mssql|SMB|RDP|Postgres|SSH|oracle|SMB2-shares)(:|\s).*)', i, re.I)
        n_p = re.findall(r'((ftp|mysql|mssql|SMB|RDP|Postgres|SSH|oracle|SMB2-shares)(:|\s).*)', i, re.I)
        rd = re.findall(r'((redis|Mongodb)(:|\s).*)', i, re.I)
        n_rd = re.findall(r'((redis|Mongodb)(:|\s).*)', i, re.I)
        mc = re.findall(r"((Memcached)(:|\s).*)", i, re.I)

        if len(p) != 0 and p[0][-1] == ":":
            p1 = list(p)

            all = p1[0][0].split(":")
            try:
                passwd = all[3]
            except:
                passwd = []
                pass
            server = all[0]
            port = all[2]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", str(all))
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

        if len(n_p) != 0 and ':' in n_p[0][0]:
            p2 = list(n_p)

            all2 = p2[0][0].split(":")
            try:
                passwd = all2[2]
            except:
                passwd = []
                pass
            server = p2[0][1]
            port = all2[1]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", str(all2))
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)



        if len(rd) != 0 and len(rd[0][0].split(" ")) == 2:
            rd1 = list(rd)

            rd_all = rd1[0][0].split(" ")
            passwd = rd_all[-1]
            server = rd1[0][1]
            port = (rd_all[0].split(":"))[2]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", rd1[0][0])
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

        if len(n_rd) != 0 and (n_rd[0][0].split(" "))[0].lower() in ["redis", "mongodb"]:
            rd2 = list(n_rd)

            rd_all2 = rd2[0][0].split(" ")
            passwd = rd_all2[2]
            server = rd_all2[0]
            try:
                port = (rd_all2[1].split(":"))[1]
            except IndexError:
                continue
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", rd2[0][0])
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)


        if len(mc) != 0:
            mc1 = list(mc)

            mc_all = mc1[0][0].split(" ")
            passwd = mc_all[2]
            server = mc_all[0]
            port = (mc_all[1].split(":"))[-1]
            ip = re.findall(r"\d+\.\d+\.\d+\.\d+", mc1[0][0])
            ip.append(port)
            ip.append(server)
            ip.append(passwd)
            sheetList.append(ip)

    OutPut('WeakPasswd', sheetList)


#输出指纹信息
def FingerOut(datastr):

    sheetList = [['url', 'finger']]

    for i in datastr:
        p = re.findall(r'.*InfoScan.*', i)


        if len(p) != 0:
            p1 = list(p)
            for u in p1:
                url = re.findall(r'http[^\s]+', u)
                finger = u.split(url[0])[-1].strip()
                url.append(finger)
                sheetList.append(url)

    OutPut('Finger', sheetList)

#输出netinfo信息
def NetInfo(datalist):

    sheetList = [['Ip', 'Netinfo', 'NetBios']]
    pattern = r'(.*NetInfo.*\n.*(\n.*\[->].*)+(\n.*NetBios.*)?)'

    info_n = re.findall(pattern, datalist)

    for i in info_n:
        ip = re.findall(r'\[\*](\d+\.\d+\.\d+\.\d+)', i[0])
        netinfo_get = re.findall(r'((\n?.*\[->].*)+)', i[0])
        netinfo = netinfo_get[0][0]
        if i[-1] == '':
            netbios = ''
        else:
            netbios_get = re.findall(r'.*NetBios.*', i[0])
            netbios = netbios_get[0]
        ip.append(netinfo)
        ip.append(netbios)
        sheetList.append(ip)

    OutPut('NetInfo', sheetList)

#表格输出整理
def OutPut(sheetname,sheetList):

    sheetName = wb.create_sheet(sheetname)



    input_filename = sys.argv[1].split(".txt")[0]

    print(f"[*]{input_filename}.txt读取成功，{input_filename}_{sheetname}.txt文件生成中······\n")

    delimiter = "================{}================\n"
    with open(f"{input_filename}_{sheetname}.txt", 'w', encoding='utf-8') as txt_file:
        txt_file.write(delimiter.format(sheetname))
        for row in sheetList:
            txt_file.write("\t".join(map(str, row)) + "\n")

    #将列表写入sheet
    for i in sheetList:
        # 解决\x03此类特殊字符报错
        try:
            sheetName.append(i)
        except openpyxl.utils.exceptions.IllegalCharacterError:
            i[-1] = ILLEGAL_CHARACTERS_RE.sub(r'', i[-1])
            sheetName.append(i)
        except Exception as e:
            print(f"err: {e}")


    #首行格式
    for row in sheetName[f"A1:{chr(65 + len(list1[0]) - 1)}1"]:
        for cell in row:
            cell.font = Font(size=12, bold=True)


def getInput():

    if len(sys.argv) != 2:
        print("\n[*] fscan结果整理脚本，输出为.xlsx文件\n\nUsage: \n\n    python3 FscanOutput.py result.txt\n")
        exit()

    if not os.path.exists(sys.argv[1]): 
        print(f"[{sys.argv[1]}] 文件不存在")
        exit()

    return sys.argv[1]

if __name__ == "__main__":

    print(r'''

============================================================
 ______                    ____        _               _   
|  ____|                  / __ \      | |             | |  
| |__ ___  ___ __ _ _ __ | |  | |_   _| |_ _ __  _   _| |_ 
|  __/ __|/ __/ _` | '_ \| |  | | | | | __| '_ \| | | | __|
| |  \__ \ (_| (_| | | | | |__| | |_| | |_| |_) | |_| | |_ 
|_|  |___/\___\__,_|_| |_|\____/ \__,_|\__| .__/ \__,_|\__|
                                          | |              
                                          |_|           Pro       
============================================================
                                        ---By zoro123 v2.3.1
           ''')
    list1, str1 = OpenFile()

    wb = openpyxl.Workbook()
    OpenPort(list1)
    AliveIp(list1)
    Bug_ExpList(list1)
    Bug_PocList(list1)
    Oslist(list1)
    GetTitle(list1)
    GetPassword(list1)
    FingerOut(list1)
    NetInfo(str1)
    ws5 = wb["Sheet"]
    wb.remove(ws5)
    input_filename = sys.argv[1].split(".txt")[0]
    Output_xlsx = (f"%s_{time.strftime('%Y-%m-%d_%H-%M-%S', time.localtime())}.xlsx" % input_filename)
    wb.save(Output_xlsx)
    print("[+]文件处理结果如下······\n")
    New_fscanxlsx = p.load_workbook(Output_xlsx)

    print("+---------------------------------+\n")
    wt = New_fscanxlsx['OpenPort']
    print("[+++]探测存活端口共计：%s 个" % (wt.max_row-1))
    wt = New_fscanxlsx['AliveIp']
    print("[+++]探测存活IP段共计：%s 个" % (wt.max_row - 1))
    wt1 = New_fscanxlsx['Bug_ExpList']
    print("[+++]Exp可利用漏洞共计：%s 个" % (wt1.max_row-1))
    wt2 = New_fscanxlsx['Bug_PocList']
    print("[+++]Poc可利用漏洞共计：%s 个" % (wt2.max_row-1))
    wt3 = New_fscanxlsx['OsList']
    print("[+++]成功识别操作系统共计：%s 个" % (wt3.max_row-1))
    wt4 = New_fscanxlsx['Title']
    print("[+++]成功探测Web服务共计：%s 条" % (wt4.max_row-1))
    wt5 = New_fscanxlsx['WeakPasswd']
    print("[+++]成功破解账号密码共计：%s 个" % (wt5.max_row-1))
    wt6 = New_fscanxlsx['Finger']
    print("[+++]成功识别指纹共计：%s 个" % (wt6.max_row-1))
    wt7 = New_fscanxlsx['NetInfo']
    print("[+++]成功识别NetInfo共计：%s 个\n" % (wt7.max_row - 1))
    print("+---------------------------------+\n")

    print('[+]结果已经整理输出至 -- %s -- 文件所在目录！\n' % sys.argv[1])
    print('--> 表格文件名为：%s\n' % Output_xlsx)
    print(f'--> 各个模块处理文件名为：{input_filename}_sheetName.txt')










